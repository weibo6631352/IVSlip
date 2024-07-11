import os
import openpyxl
import shutil


def save_to_excel(filepath, name, gender, age, drugs, date_str, price):
    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    # 拷贝Excel模板
    template_path = './输液单.xlsx'
    shutil.copy(template_path, filepath)

    workbook = openpyxl.load_workbook(filepath)

    # 更新打印页
    print_sheet = workbook['打印页']
    print_sheet['A4'] = '  1# 姓名：' + name
    print_sheet['B4'] = '性别：' + str(gender)
    print_sheet['C4'] = '年龄：' + str(age)

    print_sheet['A7'] = '  2# 姓名：' + name
    print_sheet['B7'] = '性别：' + str(gender)
    print_sheet['C7'] = '年龄：' + str(age)

    print_sheet['A10'] = '  3# 姓名：' + name
    print_sheet['B10'] = '性别：' + str(gender)
    print_sheet['C10'] = '年龄：' + str(age)

    print_sheet['A13'] = '  4# 姓名：' + name
    print_sheet['B13'] = '性别：' + str(gender)
    print_sheet['C13'] = '年龄：' + str(age)

    print_sheet['B15'] = '日期:' + date_str
    print_sheet['A15'] = '  药费: ￥' + str(price)  # 更新总计价格

    # 计算总行高
    total_height = 0
    for row in range(5, 16, 3):
        if print_sheet.row_dimensions[row].height is not None:
            total_height += print_sheet.row_dimensions[row].height
        else:
            total_height += 15  # 默认行高为15
    # 初始化药品数量列表
    drug_counts = []

    # 遍历每个药品列表
    for drug_list in drugs:
        # 过滤掉空字符串，计算药品数量
        valid_drug_count = len([drug for drug in drug_list if drug])
        # 确保药品数量至少为1
        drug_count = valid_drug_count + 1
        # 将计算结果添加到药品数量列表中
        drug_counts.append(drug_count)

    # 计算药品总数量
    total_drug_count = sum(drug_counts)

    # 动态调整行高
    for i, count in enumerate(drug_counts):
        row = 5 + i * 3
        print_sheet.row_dimensions[row].height = total_height * (count / total_drug_count)

    # 填充药品详情
    for i, drug_list in enumerate(drugs):
        details = '\n'.join([f'{drug}' for drug in drug_list if drug])  # 使用适当的空格替代\t
        if details:
            details += "\n" + "qdivgtt"
        cell = f'A{5 + i * 3}'


        print_sheet[cell] = details



    # 更新数据页
    data_sheet = workbook['数据页']
    data_sheet['B1'] = '姓名'
    data_sheet['C1'] = name
    data_sheet['B2'] = '性别'
    data_sheet['C2'] = gender
    data_sheet['B3'] = '年龄'
    data_sheet['C3'] = age
    data_sheet['B4'] = '打印日期'
    data_sheet['C4'] = date_str
    data_sheet['B5'] = '总计价格'
    data_sheet['C5'] = price

    for i, drug_list in enumerate(drugs):
        col_start = 2 + i * 3  # 每联开始的列，间隔3列
        data_sheet.cell(row=6, column=col_start).value = f'第{i + 1}联'
        for j, drug in enumerate(drug_list):
            data_sheet.cell(row=7 + j, column=col_start).value = drug

    workbook.save(filepath)
    workbook.close()

def load_from_excel(filepath):
    workbook = openpyxl.load_workbook(filepath)
    data_sheet = workbook['数据页']
    name = data_sheet.cell(row=1, column=3).value
    gender = data_sheet.cell(row=2, column=3).value
    age = data_sheet.cell(row=3, column=3).value
    date_str = data_sheet.cell(row=4, column=3).value
    price = data_sheet.cell(row=5, column=3).value
    drugs = [[] for _ in range(4)]
    for i in range(4):
        col_start = 2 + i * 3  # 每联开始的列，间隔3列
        row = 7
        for _ in range(50):
            drug = data_sheet.cell(row=row, column=col_start).value
            if drug:
                drugs[i].append(drug)
            row += 1
    return name, gender, age, drugs, price
